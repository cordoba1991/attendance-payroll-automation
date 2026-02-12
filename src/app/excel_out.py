from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime

def export_eventos_xlsx(path_out: str, rows: list[tuple]):
    """
    rows: (pin, empleado, fecha, hora, estado)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"

    ws.append(["ID", "Empleado", "Fecha", "Hora", "Estado"])
    for r in rows:
        ws.append(list(r))

    # estilo simple
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center")

    wb.save(path_out)

def export_resumen_xlsx(path_out: str, quincena_rows: list[list], diario_rows: list[list],
                        write_resumen, write_diario):
    """
    write_resumen(ws, quincena_rows) y write_diario(ws, diario_rows) los provee payroll.py
    (aquí no duplicamos lógica).
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen_quincena"
    write_resumen(ws1, quincena_rows)

    ws2 = wb.create_sheet("Detalle_diario")
    write_diario(ws2, diario_rows)

    wb.save(path_out)

def load_demo_events(path_xlsx: str) -> list[dict]:
    """
    Lee data/sample_events.xlsx con columnas:
    Nombre | Fecha | Hora | Estado
    """
    wb = load_workbook(path_xlsx)
    ws = wb.active
    header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header)}

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre = row[idx["Nombre"]]
        fecha = row[idx["Fecha"]]
        hora = row[idx["Hora"]]
        estado = row[idx["Estado"]]
        out.append({"nombre": nombre, "fecha": fecha, "hora": hora, "estado": estado})
    return out
