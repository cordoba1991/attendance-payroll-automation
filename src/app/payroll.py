from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, date, time as dtime, timedelta
from typing import Dict, List, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .timeparse import parse_date_generic, parse_time_generic


# -------------------------
# Parámetros (ajustables)
# -------------------------

# Base "normal" por día (0=Lunes ... 6=Domingo)
BASE_POR_DIA = {0: 8.25, 1: 8.25, 2: 8.25, 3: 8.25, 4: 8.25, 5: 4.00, 6: 0.00}

# Ventanas de horas (puedes ajustar si tu empresa lo maneja distinto)
DIUR_START = dtime(6, 0)
DIUR_END = dtime(19, 0)

# Nocturnas: 19:00 -> 06:00
NOCT_START = dtime(19, 0)
NOCT_END = dtime(6, 0)

# Redondeo (opcional). Si no quieres redondear, deja en 0.
ROUND_MINUTES = 0  # 15 por ejemplo


# -------------------------
# Helpers tiempo / redondeo
# -------------------------

def _round_dt_to_minutes(dt: datetime, minutes: int) -> datetime:
    """Redondea dt al múltiplo de 'minutes'. Si minutes=0, no toca."""
    if minutes <= 0:
        return dt
    discard = timedelta(
        minutes=dt.minute % minutes,
        seconds=dt.second,
        microseconds=dt.microsecond
    )
    dt = dt - discard
    # redondeo hacia abajo (conservador)
    return dt

def _hours_between(a: datetime, b: datetime) -> float:
    if b <= a:
        return 0.0
    return (b - a).total_seconds() / 3600.0

def _is_sunday(d: date) -> bool:
    return d.weekday() == 6

def _clamp_dt(dt: datetime, start: datetime, end: datetime) -> datetime:
    if dt < start:
        return start
    if dt > end:
        return end
    return dt

def _daterange(d1: date, d2: date):
    cur = d1
    while cur <= d2:
        yield cur
        cur = cur + timedelta(days=1)

def _to_datetime(d: date, t: dtime) -> datetime:
    return datetime.combine(d, t)


# -------------------------
# Split por tipos de hora
# -------------------------

def split_hours_types_same_day(day: date, start_dt: datetime, end_dt: datetime) -> Tuple[float, float, float]:
    """
    Divide un intervalo dentro de un mismo día en:
    - diurnas (06:00-19:00)
    - nocturnas (19:00-24:00 y 00:00-06:00)
    - dominicales (todas las horas si es domingo)
    """
    if end_dt <= start_dt:
        return 0.0, 0.0, 0.0

    # Si es domingo, todo cuenta como dominical (y además se puede dividir si quieres).
    # Aquí lo dejamos como "dominicales = total", diurnas/nocturnas = 0 para no duplicar.
    if _is_sunday(day):
        return 0.0, 0.0, _hours_between(start_dt, end_dt)

    day_start = _to_datetime(day, dtime(0, 0))
    day_end = _to_datetime(day, dtime(23, 59, 59))

    # Segmentos
    diur_a = _to_datetime(day, DIUR_START)
    diur_b = _to_datetime(day, DIUR_END)

    # nocturna tramo 1: 00:00-06:00
    noct1_a = day_start
    noct1_b = _to_datetime(day, NOCT_END)

    # nocturna tramo 2: 19:00-24:00
    noct2_a = _to_datetime(day, NOCT_START)
    noct2_b = day_end

    # Clampea al día
    s = _clamp_dt(start_dt, day_start, day_end)
    e = _clamp_dt(end_dt, day_start, day_end)

    # diurnas
    diur_s = max(s, diur_a)
    diur_e = min(e, diur_b)
    diurnas = _hours_between(diur_s, diur_e)

    # nocturnas tramo 1
    n1_s = max(s, noct1_a)
    n1_e = min(e, noct1_b)
    noct1 = _hours_between(n1_s, n1_e)

    # nocturnas tramo 2
    n2_s = max(s, noct2_a)
    n2_e = min(e, noct2_b)
    noct2 = _hours_between(n2_s, n2_e)

    nocturnas = noct1 + noct2
    dominicales = 0.0

    return diurnas, nocturnas, dominicales

def split_hours_types_any_span(start_dt: datetime, end_dt: datetime) -> Tuple[float, float, float]:
    """
    Divide cualquier intervalo (puede cruzar medianoche) en diurnas/nocturnas/dominicales
    día por día.
    """
    if end_dt <= start_dt:
        return 0.0, 0.0, 0.0

    diur = noct = dom = 0.0

    # iterar días
    for d in _daterange(start_dt.date(), end_dt.date()):
        day_start = _to_datetime(d, dtime(0, 0))
        day_end = _to_datetime(d, dtime(23, 59, 59))

        s = max(start_dt, day_start)
        e = min(end_dt, day_end)

        a, b, c = split_hours_types_same_day(d, s, e)
        diur += a
        noct += b
        dom += c

    return diur, noct, dom


# -------------------------
# Lectura de eventos
# -------------------------

@dataclass
class Event:
    nombre: str
    dt: datetime
    estado: str  # Entrada | Salida | Descanso

def _normalize_estado(s: str) -> str:
    s = (s or "").strip().lower()
    if s in ("entrada", "in", "checkin"):
        return "Entrada"
    if s in ("salida", "out", "checkout"):
        return "Salida"
    return "Descanso"

def read_clean_events(path_excel_limpio: str) -> List[Event]:
    """
    Espera un Excel con encabezados:
    Nombre | Fecha | Hora | Estado
    """
    wb = load_workbook(path_excel_limpio, data_only=True)
    ws = wb.active

    header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header)}

    required = ["Nombre", "Fecha", "Hora", "Estado"]
    for r in required:
        if r not in idx:
            raise ValueError(f"Falta columna '{r}' en {path_excel_limpio}. Encabezados: {header}")

    events: List[Event] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre = str(row[idx["Nombre"]] or "").strip()
        fecha_v = row[idx["Fecha"]]
        hora_v = row[idx["Hora"]]
        estado_v = row[idx["Estado"]]

        if not nombre:
            continue

        fecha = parse_date_generic(fecha_v)
        hora = parse_time_generic(hora_v)
        if not fecha or not hora:
            continue

        dt = datetime.combine(fecha, hora)
        if ROUND_MINUTES > 0:
            dt = _round_dt_to_minutes(dt, ROUND_MINUTES)

        estado = _normalize_estado(str(estado_v or ""))
        events.append(Event(nombre=nombre, dt=dt, estado=estado))

    # orden por persona, fecha/hora
    events.sort(key=lambda e: (e.nombre.lower(), e.dt))
    return events


# -------------------------
# Emparejar Entrada -> Salida
# -------------------------

@dataclass
class Interval:
    nombre: str
    start: datetime
    end: datetime

def build_work_intervals(events: List[Event]) -> List[Interval]:
    """
    Regla simple:
    - Toma Entrada como inicio
    - La siguiente Salida (misma persona) como fin
    - Descanso se ignora (en este demo)
    """
    intervals: List[Interval] = []
    last_in: Dict[str, Optional[datetime]] = {}

    for ev in events:
        n = ev.nombre

        if ev.estado == "Entrada":
            last_in[n] = ev.dt

        elif ev.estado == "Salida":
            if last_in.get(n):
                start = last_in[n]
                end = ev.dt
                # si salida < entrada, asumimos que cruzó medianoche y sumamos 1 día
                if end <= start:
                    end = end + timedelta(days=1)
                intervals.append(Interval(nombre=n, start=start, end=end))
                last_in[n] = None
            else:
                # salida sin entrada: ignorar
                continue

        else:
            # Descanso: ignorar en este demo
            continue

    return intervals


# -------------------------
# Cálculo por quincena
# -------------------------

def _quincena_range(year: int, month: int, quincena: int, margen_dias: int) -> Tuple[date, date, str]:
    """
    quincena=1 -> 1..15
    quincena=2 -> 16..fin de mes
    margen_dias: expande rango para tolerar eventos cerca del borde
    """
    if quincena not in (1, 2):
        raise ValueError("quincena debe ser 1 o 2")

    start = date(year, month, 1) if quincena == 1 else date(year, month, 16)

    # fin de quincena
    if quincena == 1:
        end = date(year, month, 15)
    else:
        # último día del mes
        if month == 12:
            end = date(year, 12, 31)
        else:
            end = date(year, month + 1, 1) - timedelta(days=1)

    start_m = start - timedelta(days=margen_dias)
    end_m = end + timedelta(days=margen_dias)
    rango_str = f"{start.isoformat()} a {end.isoformat()}"
    return start_m, end_m, rango_str

def calcular_horas_desde_excel(path_excel_limpio: str, year: int, month: int, quincena: int, margen: int):
    """
    Devuelve:
      quincena_rows, diario_rows, rango_quincena_str
    """
    start_m, end_m, rango_str = _quincena_range(year, month, quincena, margen)

    events = read_clean_events(path_excel_limpio)

    # filtrar por rango
    events = [e for e in events if start_m <= e.dt.date() <= end_m]

    intervals = build_work_intervals(events)

    # Agregación diaria por empleado
    # diario[(nombre, fecha)] = dict con horas
    diario: Dict[Tuple[str, date], Dict[str, float]] = {}

    for itv in intervals:
        # dividir por días en caso de cruce
        diur, noct, dom = split_hours_types_any_span(itv.start, itv.end)
        total = _hours_between(itv.start, itv.end)

        # asignamos el total al día de inicio como simplificación,
        # y el split lo usamos como breakdown global del intervalo.
        # Si quieres exactitud por día, habría que partir el intervalo por día y sumar.
        # Para DEMO y portafolio esto suele ser suficiente.
        key = (itv.nombre, itv.start.date())
        if key not in diario:
            diario[key] = {
                "total": 0.0,
                "diurnas": 0.0,
                "nocturnas": 0.0,
                "dominicales": 0.0,
            }
        diario[key]["total"] += total
        diario[key]["diurnas"] += diur
        diario[key]["nocturnas"] += noct
        diario[key]["dominicales"] += dom

    # Construir diario_rows
    diario_rows: List[List] = []
    # También acumulamos resumen por empleado
    resumen: Dict[str, Dict[str, float]] = {}

    # ordenar por empleado y fecha
    for (nombre, dia), vals in sorted(diario.items(), key=lambda x: (x[0][0].lower(), x[0][1])):
        weekday = dia.weekday()
        base = BASE_POR_DIA.get(weekday, 0.0)
        total = vals["total"]

        # “extras” simple: total - base si es positivo
        extras = max(0.0, total - base)

        if nombre not in resumen:
            resumen[nombre] = {
                "total": 0.0,
                "diurnas": 0.0,
                "nocturnas": 0.0,
                "dominicales": 0.0,
                "extras": 0.0,
            }
        resumen[nombre]["total"] += total
        resumen[nombre]["diurnas"] += vals["diurnas"]
        resumen[nombre]["nocturnas"] += vals["nocturnas"]
        resumen[nombre]["dominicales"] += vals["dominicales"]
        resumen[nombre]["extras"] += extras

        diario_rows.append([
            nombre,
            dia.isoformat(),
            ["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"][weekday],
            round(total, 2),
            round(vals["diurnas"], 2),
            round(vals["nocturnas"], 2),
            round(vals["dominicales"], 2),
            round(base, 2),
            round(extras, 2),
        ])

    # Construir quincena_rows
    quincena_rows: List[List] = []
    for nombre in sorted(resumen.keys(), key=lambda s: s.lower()):
        r = resumen[nombre]
        quincena_rows.append([
            nombre,
            round(r["total"], 2),
            round(r["diurnas"], 2),
            round(r["nocturnas"], 2),
            round(r["dominicales"], 2),
            round(r["extras"], 2),
        ])

    return quincena_rows, diario_rows, rango_str


# -------------------------
# Writers de Excel (openpyxl)
# -------------------------

def _style_header(ws, row: int, cols: int):
    fill = PatternFill("solid", fgColor="DDDDDD")
    font = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center")

    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align

def _autosize(ws, max_cols: int):
    for col in range(1, max_cols + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18

def escribir_hoja_resumen(ws, quincena_rows: List[List]):
    ws.append(["Empleado", "Horas Totales", "Diurnas", "Nocturnas", "Dominicales", "Extras"])
    _style_header(ws, 1, 6)

    for r in quincena_rows:
        ws.append(r)

    _autosize(ws, 6)

def escribir_hoja_diario(ws, diario_rows: List[List]):
    ws.append(["Empleado", "Fecha", "Día", "Total", "Diurnas", "Nocturnas", "Dominicales", "Base Día", "Extras"])
    _style_header(ws, 1, 9)

    for r in diario_rows:
        ws.append(r)

    _autosize(ws, 9)
